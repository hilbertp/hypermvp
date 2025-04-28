"""
UNFINISHED, BUGGY VERSION OF openpyxl+duckdb-importer

ETL (Extract, Transform, Load) pipeline for Provider Data using DuckDB's native read_excel().
This module handles validation, extraction, and atomic loading of Excel data into DuckDB.
"""
import duckdb
import openpyxl
import logging
from datetime import datetime
import os

RAW_TABLE_SCHEMA = {
    "DELIVERY_DATE": "VARCHAR",
    "PRODUCT": "VARCHAR",
    "ENERGY_PRICE_[EUR/MWh]": "VARCHAR",
    "ENERGY_PRICE_PAYMENT_DIRECTION": "VARCHAR",
    "ALLOCATED_CAPACITY_[MW]": "VARCHAR",
    "NOTE": "VARCHAR",
    "source_file": "VARCHAR",
    "load_timestamp": "TIMESTAMP"
}

def create_raw_provider_table(conn, table_name="raw_provider_data"):
    """Creates the raw provider data table if it doesn't exist."""
    columns_sql = ",\n    ".join([f'"{col}" {dtype}' for col, dtype in RAW_TABLE_SCHEMA.items()])
    create_sql = f"""
    CREATE TABLE IF NOT EXISTS {table_name} (
        {columns_sql}
    );
    """
    conn.execute(create_sql)
    logging.info(f"Ensured table '{table_name}' exists.")

def get_sheet_names(excel_filepath):
    """Returns a list of sheet names in the Excel file."""
    wb = openpyxl.load_workbook(excel_filepath, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names

def validate_excel_sheet_columns(excel_filepath, sheet_name, required_columns, expected_cols=9):
    """
    Checks if the given Excel sheet contains all required columns.
    Always reads the first `expected_cols` columns to avoid missing columns due to empty cells.
    """
    wb = openpyxl.load_workbook(excel_filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    # Always read the first expected_cols columns from the first row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=expected_cols, values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    print(f"DEBUG: Headers in {os.path.basename(excel_filepath)} sheet '{sheet_name}': {headers}")
    wb.close()
    missing = [col for col in required_columns if col not in headers]
    if missing:
        return False, missing
    return True, []

def get_min_max_delivery_date(excel_filepath, sheet_name, expected_cols=9):
    """
    Returns min and max DELIVERY_DATE in the sheet, always reading all columns.
    Uses the same robust approach as print_excel_preview.py.
    """
    wb = openpyxl.load_workbook(excel_filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    # Always read the first expected_cols columns from the header row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=expected_cols, values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    try:
        idx = headers.index("DELIVERY_DATE")
    except ValueError:
        wb.close()
        return None, None
    dates = []
    print(f"DEBUG: First 10 data rows in '{sheet_name}':")
    for i, row in enumerate(
        ws.iter_rows(min_row=2, max_row=11, min_col=1, max_col=expected_cols, values_only=True), 1
    ):
        padded = [(str(cell) if cell is not None else "") for cell in row] + [""] * (expected_cols - len(row))
        val = padded[idx]
        print(f"Row {i}: {padded} | DELIVERY_DATE: {val}")
        if val:
            dates.append(val)
    # Now scan all rows for actual min/max
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=expected_cols, values_only=True):
        padded = [(str(cell) if cell is not None else "") for cell in row] + [""] * (expected_cols - len(row))
        val = padded[idx]
        if val:
            dates.append(val)
    wb.close()
    if not dates:
        return None, None
    return min(dates), max(dates)

def validate_all_excels_in_directory(directory, required_columns):
    """
    Validates all Excel files and sheets in a directory.
    Returns (True, (min_date, max_date, file_sheet_pairs)) if all valid,
    else (False, error_message).
    """
    excel_files = [os.path.join(directory, f) for f in os.listdir(directory) if f.lower().endswith('.xlsx')]
    if not excel_files:
        return False, "No Excel files found in directory."
    min_date, max_date = None, None
    file_sheet_pairs = []
    for excel_file in excel_files:
        try:
            sheet_names = get_sheet_names(excel_file)
            for sheet in sheet_names:
                valid, missing = validate_excel_sheet_columns(excel_file, sheet, required_columns)
                if not valid:
                    return False, f"Missing columns {missing} in {os.path.basename(excel_file)} sheet '{sheet}'"
                s_min, s_max = get_min_max_delivery_date(excel_file, sheet)
                if s_min is None or s_max is None:
                    return False, f"No DELIVERY_DATE values in {os.path.basename(excel_file)} sheet '{sheet}'"
                if (min_date is None) or (s_min < min_date):
                    min_date = s_min
                if (max_date is None) or (s_max > max_date):
                    max_date = s_max
                file_sheet_pairs.append((excel_file, sheet))
        except Exception as e:
            return False, f"Error reading {excel_file}: {e}"
    return True, (min_date, max_date, file_sheet_pairs)

def atomic_load_excels(directory, db_path, table_name="raw_provider_data"):
    """
    Loads all Excel files in a directory as a single atomic operation.
    Only commits if all files/sheets are valid and loaded.
    """
    required_columns = list(RAW_TABLE_SCHEMA.keys())[:-2]  # Exclude source_file, load_timestamp
    valid, result = validate_all_excels_in_directory(directory, required_columns)
    if not valid:
        logging.error(f"Import aborted: {result}")
        return 0

    min_date, max_date, file_sheet_pairs = result
    conn = duckdb.connect(db_path)
    create_raw_provider_table(conn, table_name)
    try:
        conn.execute("BEGIN TRANSACTION;")
        # Delete all data in date range
        conn.execute(
            f"DELETE FROM {table_name} WHERE DELIVERY_DATE BETWEEN ? AND ?",
            [min_date, max_date]
        )
        total_rows = 0
        for excel_file, sheet in file_sheet_pairs:
            # Use DuckDB's read_excel for this sheet
            sql = f"""
            INSERT INTO {table_name}
            SELECT
                DELIVERY_DATE,
                PRODUCT,
                "ENERGY_PRICE_[EUR/MWh]",
                ENERGY_PRICE_PAYMENT_DIRECTION,
                "ALLOCATED_CAPACITY_[MW]",
                NOTE,
                '{os.path.basename(excel_file)}' AS source_file,
                CURRENT_TIMESTAMP AS load_timestamp
            FROM read_excel('{excel_file}', sheet='{sheet}', header=TRUE)
            """
            res = conn.execute(sql)
            rows = res.rowcount if hasattr(res, "rowcount") else 0
            total_rows += rows
            logging.info(f"Loaded {rows} rows from {os.path.basename(excel_file)} sheet '{sheet}'")
        conn.execute("COMMIT;")
        logging.info(f"Full import for {min_date} to {max_date} completed successfully. {total_rows} rows loaded.")
        return total_rows
    except Exception as e:
        conn.execute("ROLLBACK;")
        logging.error(f"Full import failed and rolled back: {e}")
        return 0
    finally:
        conn.close()