"""
Direct DuckDB Loader for Provider Data

This module provides functions to load Excel files directly into DuckDB
without going through pandas first, significantly improving performance.
"""

import os
import logging
import time
import duckdb
import concurrent.futures
import pandas as pd
import traceback

# Fix process_file to return both filename and row count as expected
def process_file(excel_file, raw_table_name, db_path):
    """
    Load an Excel file directly into DuckDB.
    """
    import os
    import sys
    import openpyxl
    import duckdb
    import time
    from datetime import datetime
    
    file_name = os.path.basename(excel_file)
    print(f"Loading file: {file_name}")
    
    # Connect to DuckDB
    conn = duckdb.connect(db_path)
    
    # Load Excel file
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    
    # Get sheet count for reporting
    sheets = wb.sheetnames
    total_rows = 0
    
    # First, count total rows for progress reporting
    for sheet_name in sheets:
        sheet = wb[sheet_name]
        # Get an estimate of row count (may not be completely accurate in read_only mode)
        sheet_rows = sheet.max_row - 1  # Subtract 1 for header
        total_rows += max(0, sheet_rows)
    
    # Now process each sheet
    processed_rows = 0
    for sheet_name in sheets:
        sheet = wb[sheet_name]
        
        # Skip empty sheets or sheets with just a header
        if sheet.max_row <= 1:
            continue
        
        # Get the headers from the first row
        headers = []
        first_row = next(sheet.rows)
        for cell in first_row:
            headers.append(str(cell.value).strip() if cell.value else "")
        
        # Process data rows
        batch_size = 10000
        batch = []
        
        # Start batch processing of rows
        for i, row in enumerate(sheet.rows):
            if i == 0:  # Skip header row
                continue
                
            # Process this row
            row_data = []
            for cell in row:
                if cell.value is None:
                    row_data.append("")
                else:
                    row_data.append(str(cell.value).strip())
            
            # Add source file and timestamp
            row_data.append(file_name)
            row_data.append(datetime.now().isoformat())
            
            batch.append(row_data)
            
            # Insert batch when it reaches the specified size
            if len(batch) >= batch_size:
                conn.execute(f"INSERT INTO {raw_table_name} VALUES {','.join(['(' + ','.join(['?' for _ in range(len(row_data))]) + ')' for _ in batch])}", [item for row in batch for item in row])
                processed_rows += len(batch)
                batch = []
                
                # Print progress update - this is the key line for the progress bar
                print(f"Processed {processed_rows:,} out of {total_rows:,} rows")
                sys.stdout.flush()
        
        # Insert any remaining rows
        if batch:
            conn.execute(f"INSERT INTO {raw_table_name} VALUES {','.join(['(' + ','.join(['?' for _ in range(len(row_data))]) + ')' for _ in batch])}", [item for row in batch for item in row])
            processed_rows += len(batch)
            
            # Print progress update for remaining rows
            print(f"Processed {processed_rows:,} out of {total_rows:,} rows")
            sys.stdout.flush()
    
    # Close the workbook
    wb.close()
    conn.close()
    
    # Final progress update
    print(f"Processed {processed_rows:,} out of {total_rows:,} rows")
    sys.stdout.flush()
    
    # Return both filename and row count to match what load_excel_to_duckdb_parallel expects
    return file_name, processed_rows

def load_excel_to_duckdb_parallel(excel_files, db_path, raw_table_name="raw_provider_data"):
    """
    Process multiple Excel files concurrently.
    """
    if not excel_files:
        logging.warning("No Excel files provided for loading")
        return 0, 0, []

    # Connect once to create the raw table if necessary
    con = duckdb.connect(db_path)
    con.execute(f"""
        CREATE TABLE IF NOT EXISTS {raw_table_name} (
            DELIVERY_DATE VARCHAR,
            PRODUCT VARCHAR,
            "ENERGY_PRICE_[EUR/MWh]" VARCHAR,
            ENERGY_PRICE_PAYMENT_DIRECTION VARCHAR,
            "ALLOCATED_CAPACITY_[MW]" VARCHAR,
            NOTE VARCHAR,
            source_file VARCHAR,
            load_timestamp TIMESTAMP
        )
    """)
    con.close()
    
    total_rows_loaded = 0
    processed_files = []
    
    # Use a process pool to avoid thread-safety issues with DuckDB connections
    with concurrent.futures.ProcessPoolExecutor() as executor:
        futures = {executor.submit(process_file, file, raw_table_name, db_path): file for file in excel_files}
        for future in concurrent.futures.as_completed(futures):
            file_name, rows = future.result()
            if rows > 0:
                processed_files.append(file_name)
                total_rows_loaded += rows
    
    success_count = len(processed_files)
    return success_count, total_rows_loaded, processed_files

# Add backward compatibility alias for legacy imports:
load_excel_to_duckdb = load_excel_to_duckdb_parallel

# Import functions from provider_db_cleaner.py for backward compatibility
# This ensures existing code that uses provider_loader.py still works
from hypermvp.provider.provider_db_cleaner import (
    analyze_raw_provider_data,
    update_provider_data_in_db,
    clean_provider_data_in_db
)