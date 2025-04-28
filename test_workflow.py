#!/usr/bin/env python3
import os
import logging
import duckdb
from tqdm import tqdm
import openpyxl
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# Define paths
EXCEL_FILE = "/home/philly/hypermvp/data/01_raw/provider/REDUCED_RESULT_LIST_ANONYM_ENERGY_MARKET_aFRR_DE_2024-09-01_11_two_sheets.xlsx"
TEST_DB_PATH = "/home/philly/hypermvp/data/04_database/test_two_sheets.duckdb"

def import_excel_direct(excel_file, db_path, table_name="raw_provider_data"):
    """Import Excel directly to DuckDB without using the problematic functions"""
    
    # Make sure the directory exists
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    
    # Remove test database if it exists
    if os.path.exists(db_path):
        logging.info(f"Removing existing database: {db_path}")
        os.remove(db_path)
    
    # Create a fresh database connection and table
    conn = duckdb.connect(db_path)
    conn.execute(f"""
        CREATE TABLE {table_name} (
            DELIVERY_DATE VARCHAR,
            PRODUCT VARCHAR,
            "ENERGY_PRICE_[EUR/MWh]" VARCHAR,
            ENERGY_PRICE_PAYMENT_DIRECTION VARCHAR,
            "ALLOCATED_CAPACITY_[MW]" VARCHAR, 
            NOTE VARCHAR,
            source_file VARCHAR,
            load_timestamp TIMESTAMP
        )
    """)
    
    file_name = os.path.basename(excel_file)
    logging.info(f"Loading Excel file: {file_name}")
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    
    total_rows = 0
    processed_rows = 0
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        row_count = sheet.max_row - 1  # Exclude header
        if row_count <= 0:
            continue
            
        logging.info(f"Processing sheet: {sheet_name} with {row_count} rows")
        total_rows += row_count
        
        # Get headers (first row) - this is critical for mapping columns correctly
        first_row = next(sheet.rows)
        headers = []
        for cell in first_row:
            value = "" if cell.value is None else str(cell.value).strip()
            headers.append(value)
            
        # Print headers to debug
        logging.info(f"Headers: {headers}")
        
        # Define the expected columns in our table
        expected_columns = [
            "DELIVERY_DATE", 
            "PRODUCT", 
            "ENERGY_PRICE_[EUR/MWh]", 
            "ENERGY_PRICE_PAYMENT_DIRECTION", 
            "ALLOCATED_CAPACITY_[MW]", 
            "NOTE"
        ]
        
        # Create mapping from Excel columns to our table columns
        col_indices = {}
        for col_name in expected_columns:
            try:
                col_indices[col_name] = headers.index(col_name)
            except ValueError:
                logging.warning(f"Column {col_name} not found in Excel headers!")
                col_indices[col_name] = None
        
        # Process data rows with progress bar
        with tqdm(total=row_count, desc=f"Sheet {sheet_name}", unit="rows") as pbar:
            batch_size = 100  # Small for testing
            batch = []
            
            for i, row in enumerate(sheet.rows):
                if i == 0:  # Skip header
                    continue
                
                # Process row - ONLY take the columns we want in the exact order of our table
                row_data = []
                for col_name in expected_columns:
                    idx = col_indices[col_name]
                    if idx is not None and idx < len(row):
                        cell = row[idx]
                        value = "" if cell.value is None else str(cell.value).strip()
                        row_data.append(value)
                    else:
                        row_data.append("") # Add empty string if column not found
                
                # Add metadata columns
                row_data.append(file_name)  # source_file
                row_data.append(datetime.now().isoformat())  # load_timestamp
                
                batch.append(row_data)
                
                # Insert batch when it reaches specified size
                if len(batch) >= batch_size:
                    # Prepare SQL for batch insert with placeholders
                    placeholders = ",".join(["(" + ",".join(["?" for _ in range(len(expected_columns) + 2)]) + ")" for _ in batch])
                    values_flat = [val for row in batch for val in row]
                    
                    # Execute the insert
                    conn.execute(f"INSERT INTO {table_name} VALUES {placeholders}", values_flat)
                    processed_rows += len(batch)
                    pbar.update(len(batch))
                    batch = []
                
            # Insert any remaining rows
            if batch:
                placeholders = ",".join(["(" + ",".join(["?" for _ in range(len(expected_columns) + 2)]) + ")" for _ in batch])
                values_flat = [val for row in batch for val in row]
                conn.execute(f"INSERT INTO {table_name} VALUES {placeholders}", values_flat)
                processed_rows += len(batch)
                pbar.update(len(batch))
    
    # Close and check result
    wb.close()
    row_count = conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
    logging.info(f"Successfully imported {row_count} rows to {db_path}")
    
    # Show some basic stats about the imported data
    if row_count > 0:
        date_stats = conn.execute(f"""
            SELECT 
                MIN(DELIVERY_DATE) as min_date, 
                MAX(DELIVERY_DATE) as max_date
            FROM {table_name}
        """).fetchone()
        logging.info(f"Date range: {date_stats[0]} to {date_stats[1]}")
    
    conn.close()
    return row_count

if __name__ == "__main__":
    import_excel_direct(EXCEL_FILE, TEST_DB_PATH)
    logging.info("Test import completed successfully")