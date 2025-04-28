"""
Utility script to print the first 10 rows (including all columns) of the first sheet
in a given Excel file, tab-separated and formatted as lines. This helps debug header and data row detection issues.
"""

import openpyxl
import sys
import os

def print_excel_preview(excel_path, sheet_name=None, num_rows=10, expected_cols=9):
    """
    Prints the first `num_rows` lines of the given Excel sheet, always showing `expected_cols` columns.
    """
    if not os.path.exists(excel_path):
        print(f"File not found: {excel_path}")
        return

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    print(f"Sheet: {sheet_name}")
    print(f"Printing the first {num_rows} rows (including header if present):\n")
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=num_rows+1, min_col=1, max_col=expected_cols, values_only=True), 1
    ):
        padded = [(str(cell) if cell is not None else "") for cell in row]
        print("\t".join(padded))
    wb.close()

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 print_excel_preview.py <excel_file_path> [sheet_name]")
        sys.exit(1)
    excel_path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
    print_excel_preview(excel_path, sheet_name)