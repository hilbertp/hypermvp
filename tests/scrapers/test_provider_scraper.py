# tests/scrapers/test_provider_scraper.py

import pytest
from unittest.mock import patch, MagicMock
from pathlib import Path
import responses
import io
import zipfile
import shutil

from hypermvp.scrapers.provider_scraper import ProviderScraper

class TestProviderScraper:
    
    def setup_method(self):
        # Create test directory
        Path("test_output").mkdir(exist_ok=True)
        self.scraper = ProviderScraper(output_dir="test_output")
    
    def teardown_method(self):
        # Clean up any test files
        shutil.rmtree("test_output", ignore_errors=True)
    
    @responses.activate
    def test_download_monthly_data(self):
        # Mock the API response
        year, month = 2024, 9
        file_name = f"RESULT_LIST_ANONYM_ENERGY_MARKET_aFRR_DE_2024-09-01_2024-09-30.xlsx.zip"
        excel_file_name = file_name.replace('.zip', '')
        
        # Get the API URL format from the config
        from hypermvp.scrapers.scraper_config import PROVIDER_CONFIG
        api_url = f"{PROVIDER_CONFIG['api_url']}/download/tenders/files/{file_name}"
        
        # Create a small fake zip file as response content
        from openpyxl import Workbook
        
        # Create a small Excel file
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "DELIVERY_DATE"
        ws['B1'] = "PRODUCT"
        ws['C1'] = "ENERGY_PRICE_[EUR/MWh]"
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Put the Excel file in a ZIP
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zf:
            zf.writestr(excel_file_name, excel_buffer.getvalue())
        
        # Add the mocked response
        responses.add(
            responses.GET,
            api_url,
            body=zip_buffer.getvalue(),
            status=200,
            content_type='application/octet-stream',
            headers={'Content-Disposition': f'attachment; filename="{file_name}"'}
        )
        
        # Patch both save_response_to_file and _extract_zip methods to actually create files
        with patch.object(self.scraper, 'save_response_to_file') as mock_save:
            with patch.object(self.scraper, '_extract_zip') as mock_extract:
                # Make the save method create the zip file
                def fake_save(response, path):
                    if isinstance(path, str):
                        path = Path(path)
                    path.parent.mkdir(parents=True, exist_ok=True)
                    with open(path, 'wb') as f:
                        f.write(zip_buffer.getvalue())
                    return path
                
                mock_save.side_effect = fake_save
                
                # Make the extract method create the Excel file
                def fake_extract(zip_path):
                    excel_path = Path("test_output") / excel_file_name
                    with open(excel_path, 'wb') as f:
                        f.write(excel_buffer.getvalue())
                    return excel_path
                
                mock_extract.side_effect = fake_extract
                
                # Call the method being tested
                result = self.scraper.download_monthly_data(year, month)
        
        # Assertions
        assert result is not None
        assert result.name == excel_file_name
    
    def test_download_error_handling(self):
        # Use direct patching of the method that's failing
        with patch.object(self.scraper, 'get_with_retry') as mock_get:
            # Setup mock to return an error response
            mock_response = MagicMock()
            mock_response.status_code = 404
            mock_response.content = b"Error content"  # Add mock content
            mock_get.return_value = mock_response
            
            # Also patch save_response_to_file to avoid file system issues
            with patch.object(self.scraper, 'save_response_to_file') as mock_save:
                # Call the method with valid dates, but the mock will return 404
                result = self.scraper.download_monthly_data(2099, 12)
                
                # Should handle the error and return None
                assert result is None
                
                # Verify save_response_to_file was called for the error
                mock_save.assert_called_once()